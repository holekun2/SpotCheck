import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


def export_to_excel(data, excel_filename, new_filename=None):
    if isinstance(data, pd.DataFrame):
        df = data
    else:
        df = pd.DataFrame(data)
    
    if new_filename:
        excel_filename = new_filename
    
    workbook = Workbook()
    worksheet = workbook.active
    
    for r, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c, value in enumerate(row, 1):
            worksheet.cell(row=r, column=c, value=str(value))
    
    apply_default_style(workbook)
    
    workbook.save(excel_filename)
    print(f"Data has been written to {excel_filename}")




    
def create_default_style():
    default_style = NamedStyle(name="default")
    default_style.font = Font(name='Calibri', size=11)
    default_style.border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    return default_style



def apply_default_style(workbook):
    default_style = create_default_style()
    workbook.add_named_style(default_style)
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                cell.style = 'default'

